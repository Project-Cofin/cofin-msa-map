from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook



class Blog:
    def __init__(self):
        # ------------------------ 바꿔야 하는 부분 -----------------------------
        # 생각해둔 키워드
        querys = ['안녕', '반가워']
        # 라벨
        label = 1
        # 파일 이름 (감정 + .xlsx)
        file_name = "tmp.xlsx"
        # -------------------------------------------------------------------------


        # 전체 샘플 문장 개수 카운트
        cnt = 1
        # 다음 문장과 같은자를 비교하기 위한 이전 문장
        before = ""
        # 페이지가 끝이 난 경우를 확인하기 위한 이전 문장
        first_string = ""
        # 페이지가 끝났을 경우를 확인하기 위한 카운트
        # 이전 문장과 두 번 이상 비슷할 경우 또는 이전 페이지와 현재 페이지의 첫 문장이 두 번 이상 비슷할 경우
        # 검색어에 의한 페이지가 끝났다고 판단
        mini_cnt = 0
        max_mini_cnt = 0
        finish_cnt = 0
        max_finish_cnt = 0
        finish_flag = False

        wb = Workbook()
        ws = wb.active
        # 검색 페이지 수
        page = 1

        for q in querys:
            # 1 페이지부터 시작
            finish_flag = False
            page = 1
            mini_cnt = 0
            max_mini_cnt = 0
            query_cnt = 0
            finish_cnt = 0
            max_finish_cnt = 0
            while (1):
                if (finish_flag):
                    break

                req = requests.get("https://kin.naver.com/search/list.nhn?query=" + q + "&page=" + str(page))
                html = req.text
                soup = BeautifulSoup(html, 'html.parser')
                mytitles = soup.select(
                    '#s_content > div.section > ul > li:nth-child(10) > dl > dt > a'
                )
                for i in range(1, 11):
                    # 문장 추출
                    if (len(soup.select('#s_content > div.section > ul > li:nth-child(%i) > dl > dt > a' % i)) == 0):
                        finish_flag = True
                        break

                    string = (soup.select('#s_content > div.section > ul > li:nth-child(%i) > dl > dt > a' % i))[0].text
                    print('[' + q + ']', string)
                    # 문장 내용
                    ws['A' + str(cnt)] = string
                    # 라벨
                    ws['B' + str(cnt)] = label
                    cnt = cnt + 1
                    query_cnt = query_cnt + 1
                    if (query_cnt % 10 == 0):
                        if (string == first_string):
                            finish_cnt = finish_cnt + 1
                            if (finish_cnt > max_finish_cnt):
                                max_finish_cnt = finish_cnt
                        else:
                            finish_cnt = 0
                        first_string = string

                    if (before == string):
                        mini_cnt = mini_cnt + 1
                        if (max_mini_cnt < mini_cnt):
                            max_mini_cnt = mini_cnt
                    else:
                        mini_cnt = 0
                    before = string
                # wb.save(file_name)
                print(file_name)
                page = page + 1
                if (max_mini_cnt > 1):
                    finish_flag = True
                if (max_finish_cnt > 1):
                    finish_flag = True
        print(cnt)

if __name__ == '__main__':
    b = Blog()